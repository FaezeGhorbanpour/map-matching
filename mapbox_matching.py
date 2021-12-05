import pandas as pd
from mapbox import MapMatcher
source = ''
tehran_ways_json = pd.read_json( "tehran_ways.json")
tehran_nodes_json = pd.read_json("tehran_nodes.json")


def load_gps_locations(gps_file_name, sheet_name):
    from openpyxl import load_workbook
    import numpy as np
    import pandas as pd

    try:
        wb = load_workbook(source + gps_file_name + '.xlsx')
    except:
        print('No such File!')
        exit()
    sheet_1 = wb[sheet_name]

    number_of_datas = sheet_1.max_row
    lats = np.zeros(number_of_datas)
    lons = np.zeros(number_of_datas)
    time = np.zeros(number_of_datas, dtype='datetime64[s]')
    locations = list()
    for i in range(1, sheet_1.max_row):
        lats[i] = sheet_1.cell(row=i + 1, column=2).value
        lons[i] = sheet_1.cell(row=i + 1, column=3).value
        converted_time = convert_time_to_datetime(sheet_1.cell(row=i + 1, column=4).value)
        time[i] = pd.to_datetime(converted_time)
        locations.append([lats[i], lons[i], time[i]])
    return locations


def convert_time_to_datetime(time_string):
    base_date = '01-01-2000'
    segments = time_string.split(':')
    time = '' + str(int(segments[0])) + ':' + str(int(segments[1])) + ':' + str(int(segments[2]))
    return base_date + ' ' + time


def convert_times_to_string(times):
    import pandas as pd
    times_string = list()

    for t in times:
        ts = pd.to_datetime(str(t))
        d = ts.strftime("%Y-%m-%dT%TZ")
        times_string.append(d)
    return times_string





def mapbox(locations):
    # mapbox

    YOUR_ACCESS_TOKEN = 'pk.eyJ1IjoiZmFlemUiLCJhIjoiY2prcDZlcm84MmlmOTNxcDBqZnY2cTAwZiJ9.1NGnObDlHtykCUGeUkcwvA'

    service = MapMatcher(access_token=YOUR_ACCESS_TOKEN)
    times_string = convert_times_to_string([i[2] for i in locations])
    gps_point_group = list()
    all_matched_points = list()
    for hundred_points in range(0, len(locations), 100):
        line = {
            "gps_precision": 10,
            "type": "Feature",
            "properties": {
                "coordTimes": times_string[hundred_points:hundred_points + 100],
                "overview": "full",
                "tidy": True,
                "steps": True,
                "annotations": True,
            },
            "geometry": {
                "type": "LineString",
                "coordinates": [[i[1], i[0]] for i in locations[hundred_points:hundred_points + 100]]
            }
        }
        response = service.match(line, profile='mapbox.driving')
        data = response.json()
        if 'features' in data:
            for feature in data['features']:
                properties = feature["properties"]
                confidence = properties['confidence']
                distance = properties['distance']
                duration = properties['duration']
                matched_points = properties['matchedPoints']
                print("Points of route:")
                print(feature['geometry']['coordinates'])
                gps_point_group.append(feature['geometry']['coordinates'])
                print('Matched points:')
                print(matched_points)
                all_matched_points.append(matched_points)
                print('confidence:  ', confidence)
                print('distance:    ', distance)
                print('duration:    ', duration)
                print()
        else:
            print(data)
    return all_matched_points, gps_point_group


def show_on_map(matching_points, matching_point_groups, gps_points, file_name):
    ### mapbox
    import folium
    tehran_map = folium.Map(location=[35.6892, 51.3890],
                            zoom_start=10)

    for i in range(len(matching_points)):
        path = [[j[1], j[0]] for j in matching_point_groups[i]]
        folium.PolyLine(path, color="blue").add_to(tehran_map)

        for j in matching_points[i]:
            folium.CircleMarker([j[1], j[0]], radius=3, color='red', fill_color='red').add_to(tehran_map)

    folium.PolyLine([xy[:-1] for xy in gps_points], color="orange").add_to(tehran_map)
    tehran_map.save(file_name + '_mapbox' + '.html')


source_track = 'sources/mapbox/'
print('Welcome to mapbox matching python api')
print('Enter the name of xlsx file including gps points lat, lon and time')
xlsx_file_name = input().split('.')[0]
print('Enter the name of sheet inside xlsx file')
sheet_name = input()
locations = load_gps_locations(xlsx_file_name, sheet_name)
filename = source_track + xlsx_file_name.replace(' ', '_') + '_' + sheet_name.replace(' ', '_')
result_points, result_groups = mapbox(locations)
show_on_map(result_points, result_groups, locations, filename)
