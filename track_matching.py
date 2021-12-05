import pandas as pd

source = ''
tehran_ways_json = pd.read_json(source + "tehran_ways.json")
tehran_nodes_json = pd.read_json(source + "tehran_nodes.json")
tehran_ways = tehran_ways_json.set_index("id", drop=True)
tehran_nodes = tehran_nodes_json.set_index("id", drop=True)


def load_gps_locations(gps_file_name, sheet_name):
    from openpyxl import load_workbook
    import numpy as np
    import pandas as pd

    try:
        wb = load_workbook(source + gps_file_name + '.xlsx')
    except:
        print('No such File!')
        exit()
    sheet_1 = wb[sheet_name]

    number_of_datas = sheet_1.max_row
    lats = np.zeros(number_of_datas)
    lons = np.zeros(number_of_datas)
    time = np.zeros(number_of_datas, dtype='datetime64[s]')
    locations = list()
    for i in range(1, sheet_1.max_row):
        lats[i] = sheet_1.cell(row=i + 1, column=2).value
        lons[i] = sheet_1.cell(row=i + 1, column=3).value
        converted_time = convert_time(sheet_1.cell(row=i + 1, column=4).value)
        time[i] = pd.to_datetime(converted_time)
        locations.append([lats[i], lons[i], time[i]])
    return locations


def convert_time(time_string):
    base_date = '01-01-2000'
    segments = time_string.split(':')
    time = '' + str(int(segments[0])) + ':' + str(int(segments[1])) + ':' + str(int(segments[2]))
    return base_date + ' ' + time


def write_query_in_file(locations, file_name):
    context = ''
    for i in range(len(locations) - 1):
        context += str(i) + ','
        context += str(locations[i][1]) + ','
        context += str(locations[i][0]) + ','
        context += '"' + str(locations[i][2]) + '"\n'
    file = open(file_name + '_track' + '.csv', 'w')
    file.write(context)
    file.close()


def query(file_name):
    # import os
    import requests, json

    headers = {
        'Content-Type': 'text/csv',
        'Accept': 'application/json',
    }

    params = (
        ('app_id', 'e62c3c13'),
        ('app_key', 'da12514566c46d73a1a49575deb75738'),
    )

    data = open(file_name + '_track.csv', 'rb').read()
    response = requests.post('http://test.roadmatching.com/rest/mapmatch/', headers=headers, params=params, data=data)
    # print(response.text)
    return json.loads(response.text)


def track_matching(data):
    try:
        links = data.get('diary').get('entries')[0].get('route').get('links')
    except:
        print("Error in reading output file")
        exit()
    nd_lists = list()
    for link in links:
        link_id = link.get('id')
        tehran_ways = tehran_ways_json.loc[tehran_ways_json['id'] == link_id]
        links_nd = tehran_ways.get('nd')
        for nd in links_nd:
            if link['src'] in nd:
                start_index = nd.index(link['src'])
            else:
                start_index = 0
            if link['dst'] in nd:
                end_index = nd.index(link['dst'])
            else:
                end_index = -1
            if (start_index > end_index):
                start_index, end_index = end_index, start_index
            nd_lists.append(nd[start_index: end_index + 1])
    nd_lat_lon = list()
    for path in nd_lists:
        each_path_nd = list()
        for nd_id in path:
            nd_lan_lot = tehran_nodes_json.loc[tehran_nodes_json['id'] == nd_id]
            lat = nd_lan_lot.values.tolist()[0][1]
            lon = nd_lan_lot.values.tolist()[0][2]
            each_path_nd.append([lat, lon])
        nd_lat_lon.append(each_path_nd)
    return nd_lat_lon


def show_on_map(matching_point, gps_point, file_name):
    import folium
    tehran_map_2 = folium.Map(location=[35.6892, 51.3890],
                              zoom_start=10)
    # calculated road
    folium.Marker(location=matching_point[0][0],
                  icon=folium.Icon(color='green')).add_to(tehran_map_2)
    for path in matching_point:
        if path != []:
            folium.PolyLine(path, color="blue").add_to(tehran_map_2)
        folium.Marker(location=matching_point[-1][-1],
                      icon=folium.Icon(color='red')).add_to(tehran_map_2)

    # gps road
    folium.PolyLine(
        [xy[:-1] for xy in gps_point],
        color="orange"
    ).add_to(tehran_map_2)

    tehran_map_2.save(file_name + '_track' + '.html')


source_track = 'sources/track_matching/'
print('Welcome to track matching python api')
print('Enter the name of xlsx file including gps points lat, lon and time')
xlsx_file_name = input().split('.')[0]
print('Enter the name of sheet inside xlsx file')
sheet_name = input()
locations = load_gps_locations(xlsx_file_name, sheet_name)
filename = source_track + xlsx_file_name.replace(' ', '_') + '_' + sheet_name.replace(' ', '_')
write_query_in_file(locations, filename)
query_result = query(filename)

result_points = track_matching(query_result)
show_on_map(result_points, locations, filename)

# Sample Sensor Data
# Sheet1
