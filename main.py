from openpyxl import load_workbook
import numpy as np
import pandas as pd
from rdp import rdp
from mapbox_matching import MapMatcher

YOUR_ACCESS_TOKEN = 'pk.eyJ1IjoiZmFlemUiLCJhIjoiY2prcDZlcm84MmlmOTNxcDBqZnY2cTAwZiJ9.1NGnObDlHtykCUGeUkcwvA'


def convert_time(time_string):
    base_date = '01-01-2000'
    segments = time_string.split(':')
    time = '' + str(int(segments[0])) + ':' + str(int(segments[1])) + ':' + str(int(segments[2]))
    return base_date + ' ' + time


# import required libraries

times_string = list()

# read  from excel file
wb = load_workbook('Sample Sensor Data.xlsx')

sheet_1 = wb['Sheet10']

number_of_datas = sheet_1.max_row
lats = np.zeros(number_of_datas)
lons = np.zeros(number_of_datas)
times = np.zeros(number_of_datas, dtype='datetime64[s]')

locations = list()
for i in range(1, sheet_1.max_row):
    lats[i] = sheet_1.cell(row=i + 1, column=2).value
    lons[i] = sheet_1.cell(row=i + 1, column=3).value
    converted_time = convert_time(sheet_1.cell(row=i + 1, column=4).value)
    times[i] = pd.to_datetime(converted_time)
    locations.append([lats[i], lons[i], times[i]])
lats = np.delete(lats, 0)
lons = np.delete(lons, 0)
times = np.delete(times, 0)

redundant_indicators = list()
for t in range(1, len(times)):
    if times[t] <= times[t - 1]:
        redundant_indicators.append(t)
        print('redundant indicators1', t, times[t], times[t-1])

for rt in redundant_indicators:
    times = np.delete(times, rt)
    lats = np.delete(lats, rt)
    lons = np.delete(lons, rt)
    del locations[rt]

for t in range(1, len(times)):
    if times[t] <= times[t - 1]:
        redundant_indicators.append(t)

for t in times:
    ts = pd.to_datetime(str(t))
    d = ts.strftime("%Y-%m-%dT%TZ")
    times_string.append(d)



felan = rdp([xy[0:2] for xy in locations])
print('felan')
print(felan)
print(len(felan), len(locations))
print()


service = MapMatcher(access_token=YOUR_ACCESS_TOKEN)

for hundred_points in range(0, len(locations), 100):
    line = {
        "gps_precision": 10,
        "type": "Feature",
        "properties": {
            "coordTimes": times_string[hundred_points:hundred_points + 100]
        },
        "geometry": {
            "type": "LineString",
            "coordinates": [[i[1], i[0]] for i in locations[hundred_points:hundred_points + 100]]
        }
    }
    response = service.match(line, profile='mapbox.driving')
    data = response.json()
    if 'features' in data:
        for feature in data['features']:
            properties = feature["properties"]
            confidence = properties['confidence']
            distance = properties['distance']
            duration = properties['duration']
            matched_points = properties['matchedPoints']
            print("Points of route:")
            print(feature['geometry']['coordinates'])
            print('Matched points:')
            print(matched_points)
            print('confidence:  ', confidence)
            print('distance:    ', distance)
            print('duration:    ', duration)
            print()
